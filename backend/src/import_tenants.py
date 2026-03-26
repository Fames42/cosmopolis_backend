"""
Import tenants from info.xlsx into the database.

Usage (inside Docker):
    docker compose run --rm backend python -m src.import_tenants
"""

import re
import os
from datetime import datetime
from pathlib import Path

import openpyxl

from src.database import SessionLocal, engine
from src.models import Base, Building, Conversation, Message, Tenant, Ticket, TicketNote, User


# --- Cyrillic → Latin normalization ---
CYRILLIC_MAP = str.maketrans("АВСЕНКМОРТХаевскморх", "ABCEHKMOPTXaeвckмopx")


def normalize_text(s: str) -> str:
    """Replace Cyrillic look-alikes with Latin equivalents."""
    return s.translate(CYRILLIC_MAP)


# Known building prefixes (longest first for greedy matching)
KNOWN_BUILDINGS = sorted([
    "ARMAN VILLA",
    "ESENTAI APARTMENTS A",
    "ESENTAI APARTMENTS B",
    "IVANILOVA",
    "SAMAL TOWERS",
    "SAMAL DELUXE",
    "SOLNECHNAYA DOLINA",
    "STOLICHNIY",
    "ZHAMAKAYEV HOUSING",
    "DOSTYK",
    "EXCLUSIVE TIME",
    "SNEGINA",
    "AFD PLAZA",
    "ORION",
    "TAU SHATYR",
], key=len, reverse=True)

# Additional buildings with non-Latin names
EXTRA_BUILDINGS = [
    "ЖК Амир",
    "ЖК Royal",
]

PHONE_RE = re.compile(r"[\+]?[\d][\d\s\-\(\)]{8,}")


def split_building_apartment(obj: str) -> tuple[str, str]:
    """Split 'ESENTAI APARTMENTS A 8C' -> ('ESENTAI APARTMENTS A', '8C')."""
    obj_upper = normalize_text(obj.upper().strip())

    for prefix in KNOWN_BUILDINGS:
        if obj_upper.startswith(prefix):
            apt = obj[len(prefix):].strip()
            return prefix, apt

    # Check non-Latin buildings
    for prefix in EXTRA_BUILDINGS:
        if obj.startswith(prefix):
            apt = obj[len(prefix):].strip()
            return prefix, apt

    # Fallback: entire string is building, no apartment
    return obj.strip(), ""


def extract_phone(text: str) -> str:
    """Extract first phone number from a contact string."""
    if not text:
        return ""
    match = PHONE_RE.search(text)
    if match:
        return match.group(0).strip()
    return ""


def extract_name_from_contact(text: str) -> str:
    """Extract name portion before the phone number."""
    if not text:
        return ""
    match = PHONE_RE.search(text)
    if match:
        name = text[:match.start()].strip().rstrip("+").strip()
        return name
    return text.strip()


def normalize_category(raw: str) -> str:
    """Normalize category: Cyrillic А→A, В→B, С→C, etc."""
    CYRILLIC_TO_LATIN = {"А": "A", "В": "B", "С": "C", "а": "A", "в": "B", "с": "C"}
    raw = raw.strip()
    # Single-char Cyrillic → Latin
    if raw in CYRILLIC_TO_LATIN:
        return CYRILLIC_TO_LATIN[raw]
    normalized = normalize_text(raw).upper()
    if normalized in ("A", "B", "C"):
        return normalized
    if "не обслуживаем" in raw.lower():
        return "no_service"
    if normalized in ("-", ""):
        return ""
    return raw


def format_date(dt) -> str:
    """Format datetime to 'YYYY-MM-DD' string."""
    if isinstance(dt, datetime):
        return dt.strftime("%Y-%m-%d")
    if dt:
        return str(dt)
    return ""


def compute_lease_duration(start, end) -> str:
    """Compute lease duration as 'X months' from start and end dates."""
    if not isinstance(start, datetime) or not isinstance(end, datetime):
        if end:
            return f"до {format_date(end)}"
        return ""
    delta_months = (end.year - start.year) * 12 + (end.month - start.month)
    if delta_months > 0:
        return f"{delta_months} мес. (до {format_date(end)})"
    return f"до {format_date(end)}"


def clean_numeric(val) -> str:
    """Convert numeric xlsx value to clean string (strip .0 from floats)."""
    if val is None:
        return ""
    if isinstance(val, float):
        if val == int(val):
            return str(int(val))
        return str(val)
    if isinstance(val, int):
        return str(val)
    return str(val).strip()


def parse_xlsx(filepath: str) -> list[dict]:
    """Parse info.xlsx and return normalized tenant dicts."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb[wb.sheetnames[0]]

    tenants = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        obj = row[1]
        if not obj or not str(obj).strip():
            continue

        obj = str(obj).strip()
        building_name, apartment = split_building_apartment(obj)

        # New column layout (columns C-H are building detail fields)
        address = str(row[2]).strip() if row[2] else ""
        house_number = clean_numeric(row[3])
        legal_number = clean_numeric(row[4])
        floor = clean_numeric(row[5])
        block = clean_numeric(row[6])
        actual_number = clean_numeric(row[7])

        lease_start = row[8]
        lease_end = row[9]
        category_raw = str(row[10]).strip() if row[10] else ""
        client_name = str(row[11]).strip() if row[11] else ""
        company = str(row[12]).strip() if row[12] else ""
        primary_contact = str(row[13]).strip() if row[13] else ""
        extra_contact = str(row[14]).strip() if row[14] else ""

        # Normalize
        category = normalize_category(category_raw)
        phone = extract_phone(primary_contact)

        # Fallback: try extracting phone from extra contact
        if not phone and extra_contact:
            phone = extract_phone(extra_contact)

        # Tenant name: prefer client column, fallback to contact name
        name = client_name if client_name and client_name != "-----" else ""
        if not name:
            name = extract_name_from_contact(primary_contact)
        if not name:
            name = "Неизвестный"

        # Clean up company
        if company in ("-----", "None", ""):
            company = ""

        tenants.append({
            "building_name": building_name,
            "apartment": apartment,
            "address": address if address and address != "None" else "",
            "house_number": house_number,
            "legal_number": legal_number,
            "floor": floor,
            "block": block,
            "actual_number": actual_number,
            "name": name,
            "phone": phone,
            "lease_start_date": format_date(lease_start),
            "lease_end_date": format_date(lease_end),
            "emergency_contact": extra_contact if extra_contact and extra_contact != "None" else "",
            "notes": "",
            "category": category,
            "company": company,
            "agent_enabled": False,
        })

    return tenants


def import_tenants():
    """Main import function."""
    xlsx_path = os.path.join(os.path.dirname(__file__), "..", "info.xlsx")
    # Docker workdir is /app, so also check there
    if not Path(xlsx_path).exists():
        xlsx_path = "/app/info.xlsx"
    if not Path(xlsx_path).exists():
        print(f"ERROR: {xlsx_path} not found")
        return

    print("Parsing info.xlsx...")
    rows = parse_xlsx(xlsx_path)
    print(f"Parsed {len(rows)} tenant rows")

    db = SessionLocal()
    try:
        # Look up Zhanna as building owner
        zhanna = db.query(User).filter(User.email == "zhanna@cosmorent.kz").first()
        if not zhanna:
            print("ERROR: User zhanna@cosmorent.kz not found. Create the user first.")
            return
        owner_id = zhanna.id

        # Clear existing data (respecting FK order)
        db.query(TicketNote).delete()
        db.query(Ticket).delete()
        db.query(Message).delete()
        db.query(Conversation).delete()
        deleted_tenants = db.query(Tenant).delete()
        deleted_buildings = db.query(Building).delete()
        db.flush()
        print(f"Cleared {deleted_tenants} tenants, {deleted_buildings} buildings (and related tickets/conversations)")

        # Create one building per row and link tenant to it
        buildings_created = 0
        tenants_created = 0
        for r in rows:
            b = Building(
                name=r["building_name"],
                address=r["address"] or r["building_name"],
                house_number=r["house_number"] or None,
                legal_number=r["legal_number"] or None,
                floor=r["floor"] or None,
                block=r["block"] or None,
                actual_number=r["actual_number"] or None,
                owner_id=owner_id,
            )
            db.add(b)
            db.flush()
            buildings_created += 1

            t = Tenant(
                name=r["name"],
                phone=r["phone"],
                building_id=b.id,
                apartment=r["apartment"],
                lease_start_date=r["lease_start_date"] or None,
                lease_end_date=r["lease_end_date"] or None,
                emergency_contact=r["emergency_contact"] or None,
                notes=r["notes"] or None,
                category=r["category"] or None,
                company=r["company"] or None,
                agent_enabled=r["agent_enabled"],
            )
            db.add(t)
            tenants_created += 1

        db.commit()

        print(f"\n=== Import Summary ===")
        print(f"Buildings created: {buildings_created}")
        print(f"Tenants imported:  {tenants_created}")
        print("Done!")

    finally:
        db.close()


if __name__ == "__main__":
    import_tenants()
